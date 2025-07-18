from time import sleep

import mysql.connector
import pandas as pd
from datetime import datetime
import os
import ast

DB_CONFIG = {
    'host': 'localhost',
    'port': 3307,
    'user': 'root',
    'password': '',
    'database': 'SmartEmkWarehouseDB',
}

EXPORT_FOLDER = "/home/smartemk221/Desktop/wms/WarehouseManagementWeb/storage/app/public"
EXPORT_FOLDER_LOCAL = "/home/smartemk221/Desktop/wms/WarehouseManagementWebLocal/storage/app/public"

# === GET PENDING JOB ===
def get_pending_export_job():
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT * FROM export_jobs
        WHERE table_name IN ('orders', 'orders_logs') AND status = 'pending'
        ORDER BY id ASC
        LIMIT 1
    """)
    job = cursor.fetchone()
    cursor.close()
    conn.close()
    return job


# === UPDATE JOB STATUS ===
def update_job_status(job_id, status, percent=0):
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    cursor.execute("UPDATE export_jobs SET status=%s, percent=%s WHERE id=%s", (status, percent, job_id))
    conn.commit()
    cursor.close()
    conn.close()


# === EXPORT TO EXCEL ===
def export_order_items_to_excel(job):
    job_id = job["id"]

    # JSON benzeri search_values stringini Python dict'e çevir
    try:
        search_values = ast.literal_eval(job["search_values"])
    except Exception as e:
        print(f"[!] search_values ayrıştırma hatası: {e}")
        update_job_status(job_id, 'failed', 0)
        return

    order_id = search_values.get("order_id")
    if not order_id:
        update_job_status(job_id, 'failed', 0)
        print("order_id bulunamadı.")
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"orders_{order_id}_{timestamp}"
    if search_values.get("local_host"):
        file_path = os.path.join(EXPORT_FOLDER_LOCAL, f"orderList/{job['file_name']}")
    else:
        file_path = os.path.join(EXPORT_FOLDER, f"orderList/{job['file_name']}")

    try:
        update_job_status(job_id, 'processing', 0)

        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        # SQL sorgusu
        cursor.execute("""
            SELECT 
                o.id AS OrderItemId,
                o.order_id AS OrderId,
                o.order_sort_num AS OrderItemOrderNumber,
                s.code AS ItemCode,
                s.name AS ItemName,
                s.feature AS ItemDescription,
                s.production_date AS ItemProductionDate,
                s.weight AS ItemWeight,
                s.volume AS ItemVolume,
                GROUP_CONCAT(b.barcode) AS Barcode,
                o.orderQty AS OrderQty,
                o.pickingQty AS PickingQty,
                w.name AS PickPlace_W,
                l.name AS PickPlace_L,
                b2.name AS PickPlace_B,
                o.putawayQty AS PutawayQty,
                o.putaway_pin AS PutawayLocId,
                o.shipping_number AS ShippingNumber,
                c.id AS CurrCustomerId,
                c.name AS CurrCustomerName,
                c.post_code AS CurrCustomerPostCode,
                c.phone AS CurrCustomerPhone,
                c.email AS CurrCustomerEmail
            FROM order_items o
            LEFT JOIN current_stocks cs ON cs.id = o.curr_stk_id
            LEFT JOIN stocks s ON s.id = cs.stock_id
            LEFT JOIN barcodes b ON b.curr_stk_id = cs.id
            LEFT JOIN boxes b2 ON b2.id = cs.box_id
            LEFT JOIN locations l ON l.id = b2.location_id
            LEFT JOIN warehouses w ON w.id = l.warehouse_id
            LEFT JOIN customers c ON c.id = o.customer_id
            WHERE o.order_id = %s
            GROUP BY o.id
        """, (order_id,))

        rows = cursor.fetchall()
        df = pd.DataFrame(rows)

        os.makedirs(EXPORT_FOLDER, exist_ok=True)
        df.to_excel(file_path, index=False)

        # Güncelleme
        cursor.execute("""
            UPDATE export_jobs 
            SET status=%s, percent=%s, file_name=%s, file_path=%s 
            WHERE id = %s
        """, ('done', 100, file_name, file_path, job_id))
        conn.commit()

        cursor.close()
        conn.close()

        print(f"[✓] Export tamamlandı: {file_path}")
    except Exception as e:
        update_job_status(job_id, 'failed', 0)
        print(f"[!] Export işlemi başarısız: {e}")


def export_orders_logs_to_excel(job):
    job_id = job["id"]

    try:
        # search_values alanını dict olarak ayrıştır
        search_values = ast.literal_eval(job["search_values"])
    except Exception as e:
        print(f"[!] search_values ayrıştırma hatası: {e}")
        update_job_status(job_id, 'failed', 0)
        return

    # Tarih aralığı ve filtre kontrolü
    min_time = search_values.get("min")
    max_time = search_values.get("max")
    action = search_values.get("action")

    if not min_time or not max_time:
        update_job_status(job_id, 'failed', 0)
        print("min veya max zamanı eksik.")
        return

    try:
        # Zamanları datetime'a çevir (UTC+3 farkı düşülerek)
        min_dt = datetime.strptime(min_time, "%Y-%m-%dT%H:%M") - pd.Timedelta(hours=3)
        max_dt = datetime.strptime(max_time, "%Y-%m-%dT%H:%M") - pd.Timedelta(hours=3)
    except Exception as e:
        print(f"[!] Zaman ayrıştırma hatası: {e}")
        update_job_status(job_id, 'failed', 0)
        return

    # Zaman etiketi üret
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Gerçek dosya ismi ve geçici (tamamlanmamış) dosya ismi
    final_name = f"{job['file_name']}"
    temp_name = f"uncopleted_{timestamp}.xlsx"

    # Hedef klasörü belirle
    if search_values.get("local_host"):
        folder = os.path.join(EXPORT_FOLDER_LOCAL, "orderLog")
    else:
        folder = os.path.join(EXPORT_FOLDER, "orderLog")

    # Dosya yolları
    temp_path = os.path.join(folder, temp_name)
    final_path = os.path.join(folder, final_name)

    try:
        update_job_status(job_id, 'processing', 0)

        os.makedirs(folder, exist_ok=True)

        # Veritabanına bağlan
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        # SQL sorgusu (action varsa filtreli)
        sql = f"""
            SELECT 
                ol.id, ol.order_id, ol.order_item_id, ol.order_sort_num,
                s.code AS ItemCode, s.name AS ItemName, s.feature AS ItemDescription,
                s.production_date AS ItemProductionDate, s.weight AS ItemWeight, s.volume AS ItemVolume,
                ol.used_barcode_num AS Barcode, ol.orderQty, ol.pickingQty,
                w.name AS PickPlace_W, l.name AS PickPlace_L, b.name AS PickPlace_B,
                ol.putawayQty, ol.putaway_pin, ol.shipping_number,
                c.id AS CurrCustomerId, c.name AS CurrCustomerName, c.post_code, c.phone, c.email,
                ol.action, ol.created_at, u.name AS Created_by
            FROM orders_logs ol
            LEFT JOIN current_stocks cs ON cs.id = ol.curr_stk_id
            LEFT JOIN stocks s ON s.id = cs.stock_id
            LEFT JOIN boxes b ON b.id = cs.box_id
            LEFT JOIN locations l ON l.id = b.location_id
            LEFT JOIN warehouses w ON w.id = l.warehouse_id
            LEFT JOIN customers c ON c.id = ol.customer_id
            LEFT JOIN users u ON u.id = ol.created_by
            WHERE ol.created_at BETWEEN %s AND %s
        """
        params = [min_dt.strftime("%Y-%m-%d %H:%M:%S"), max_dt.strftime("%Y-%m-%d %H:%M:%S")]
        if action:
            sql += " AND ol.action = %s"
            params.append(action)

        cursor.execute(sql, params)
        rows = cursor.fetchall()
        df = pd.DataFrame(rows)

        # --- created_at datetime ve ElapsedSeconds hesaplama ---

        if 'created_at' in df.columns:
            try:
                # created_at'ı datetime formatına çevir
                df['created_at'] = pd.to_datetime(df['created_at'], errors='coerce')

                # ElapsedSeconds hesapla (ilk satır 0 olur)
                df['ElapsedSeconds'] = df['created_at'].diff().dt.total_seconds().fillna(0).astype(int)

                # ElapsedSeconds kolonunu created_at'ın hemen sonrasına taşı
                cols = df.columns.tolist()
                if 'ElapsedSeconds' in cols:
                    cols.remove('ElapsedSeconds')
                    created_at_index = cols.index('created_at')
                    cols.insert(created_at_index + 1, 'ElapsedSeconds')
                    df = df[cols]

            except Exception as e:
                print(f"[!] created_at işlem hatası: {e}")
                df['ElapsedSeconds'] = 0  # fallback
        else:
            print("[!] 'created_at' sütunu yok, ElapsedSeconds sıfırlandı.")
            df['ElapsedSeconds'] = 0

        # Geçici dosyaya yaz
        df.to_excel(temp_path, index=False)

        # Excel biçimlendirme
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Border, Side
        wb = load_workbook(temp_path)
        ws = wb.active

        # ElapsedSeconds sütununu bul
        elapsed_col_idx = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "ElapsedSeconds":
                elapsed_col_idx = col
                break

        # Koşullu renklendirme
        if elapsed_col_idx:
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=elapsed_col_idx)
                try:
                    value = int(cell.value)
                    if value >= 10:
                        cell.fill = red_fill
                    elif value >= 5:
                        cell.fill = orange_fill
                    else:
                        cell.fill = green_fill
                except:
                    continue

        # Süre istatistikleri
        created_at_list = df['created_at'].tolist()
        elapsed_seconds_list = df['ElapsedSeconds'].tolist()

        start_time = created_at_list[0]
        end_time = created_at_list[-1]
        total_seconds = int((end_time - start_time).total_seconds())
        total_h = total_seconds // 3600
        total_m = (total_seconds % 3600) // 60
        total_s = total_seconds % 60
        human_readable = f"{total_h} saat {total_m} dakika {total_s} saniye"

        avg_seconds = sum(elapsed_seconds_list) / len(elapsed_seconds_list)
        filtered_seconds = [s for s in elapsed_seconds_list if s < 10]
        filtered_avg_seconds = sum(filtered_seconds) / len(filtered_seconds) if filtered_seconds else 0
        filtered_total_duration = filtered_avg_seconds * len(df)
        ft_h = int(filtered_total_duration // 3600)
        ft_m = int((filtered_total_duration % 3600) // 60)
        ft_s = int(filtered_total_duration % 60)

        # Excel alt satırlara özet yaz
        footer_row = ws.max_row + 2
        ws[f"A{footer_row}"] = "Gercek"
        ws[f"B{footer_row}"] = human_readable
        ws[f"C{footer_row}"] = round(avg_seconds, 6)

        ws[f"A{footer_row + 1}"] = "Filtrelenmiş Ortalama Süre (Kırmızı değerler hariç)"
        ws[f"B{footer_row + 1}"] = f"{ft_h} saat {ft_m} dakika {ft_s} saniye"
        ws[f"C{footer_row + 1}"] = round(filtered_avg_seconds, 6)

        # Kenarlıklar
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for row in range(footer_row, footer_row + 2):
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = thin_border

        # Dosyayı kaydet
        wb.save(temp_path)

        # Geçici dosyayı gerçek isme taşı
        os.rename(temp_path, final_path)

        # Veritabanında güncelle
        cursor.execute("""
            UPDATE export_jobs 
            SET status=%s, percent=%s, file_name=%s, file_path=%s 
            WHERE id = %s
        """, ('done', 100, final_name, final_path, job_id))
        conn.commit()

        cursor.close()
        conn.close()

        print(f"[✓] Export tamamlandı: {final_path}")

    except Exception as e:
        update_job_status(job_id, 'failed', 0)
        print(f"[!] Export başarısız: {e}")

# === MAIN ===
if __name__ == "__main__":
    while 1:
        job = get_pending_export_job()
        if job:
            table = job['table_name']
            if table == 'orders':
                export_order_items_to_excel(job)
            elif table == 'orders_logs':
                export_orders_logs_to_excel(job)
            else:
                print(f"[!] Desteklenmeyen tablo: {table}")
        else:
            print("[✓] Bekleyen iş yok.")

        sleep(10);


