from time import sleep
import mysql.connector
import pandas as pd
from datetime import datetime
import os
import ast
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

from v06.job_utils import update_job_status
from v06.config import EXPORT_FOLDER_LOCAL, EXPORT_FOLDER, DB_CONFIG


class ExportOrdersLogsJob:
    def __init__(self, job):
        self.job = job
        self.job_id = job["id"]
        self.search_values = {}
        self.df = pd.DataFrame()
        self.min_dt = None
        self.max_dt = None
        self.folder = ""
        self.temp_path = ""
        self.final_path = ""
        self.elapsed_col_idx = None
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    def run(self):
        try:
            self._parse_search_values()
            self._set_time_range()
            self._prepare_paths()
            update_job_status(self.job_id, 'processing', 0)
            rows = self._fetch_data()
            self._create_dataframe(rows)
            self._calculate_elapsed_seconds()
            self._write_excel()
            self._apply_formatting()
            self._save_and_finalize()
            self._mark_done()
            print(f"[✓] Export tamamlandı: {self.final_path}")
        except Exception as e:
            update_job_status(self.job_id, 'failed', 0)
            print(f"[!] Export başarısız: {e}")

    def _parse_search_values(self):
        try:
            self.search_values = ast.literal_eval(self.job["search_values"])
        except Exception as e:
            raise ValueError(f"search_values ayrıştırma hatası: {e}")

    def _set_time_range(self):
        min_time = self.search_values.get("min")
        max_time = self.search_values.get("max")

        if not min_time or not max_time:
            raise ValueError("min veya max zamanı eksik.")

        self.min_dt = datetime.strptime(min_time, "%Y-%m-%dT%H:%M") - pd.Timedelta(hours=3)
        self.max_dt = datetime.strptime(max_time, "%Y-%m-%dT%H:%M") - pd.Timedelta(hours=3)

    def _prepare_paths(self):
        file_name = self.job['file_name']
        temp_name = f"uncopleted_{self.timestamp}.xlsx"
        is_local = self.search_values.get("local_host")

        base_folder = EXPORT_FOLDER_LOCAL if is_local else EXPORT_FOLDER
        self.folder = os.path.join(base_folder, "orderLog")
        os.makedirs(self.folder, exist_ok=True)

        self.temp_path = os.path.join(self.folder, temp_name)
        self.final_path = os.path.join(self.folder, file_name)

    def _fetch_data(self):
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)

        sql = """
            SELECT 
                ol.id, ol.order_id, ol.order_item_id, ol.order_sort_num,
                s.code AS ItemCode, s.name AS ItemName, s.feature AS ItemDescription,
                s.production_date AS ItemProductionDate, s.weight AS ItemWeight, s.volume AS ItemVolume,
                ol.used_barcode_num AS Barcode, ol.orderQty, ol.pickingQty,
                w.name AS PickPlace_W, l.name AS PickPlace_L, b.name AS PickPlace_B,
                ol.putawayQty, ol.putaway_pin, ol.shipping_number,
                c.id AS CurrCustomerId, c.name AS CurrCustomerName, c.post_code, c.phone, c.email,
                ol.action, ol.created_at, u.name AS Created_by
            FROM orders_logs ol
            LEFT JOIN current_stocks cs ON cs.id = ol.curr_stk_id
            LEFT JOIN stocks s ON s.id = cs.stock_id
            LEFT JOIN boxes b ON b.id = cs.box_id
            LEFT JOIN locations l ON l.id = b.location_id
            LEFT JOIN warehouses w ON w.id = l.warehouse_id
            LEFT JOIN customers c ON c.id = ol.customer_id
            LEFT JOIN users u ON u.id = ol.created_by
            WHERE ol.created_at BETWEEN %s AND %s
        """

        params = [self.min_dt.strftime("%Y-%m-%d %H:%M:%S"), self.max_dt.strftime("%Y-%m-%d %H:%M:%S")]
        action = self.search_values.get("action")
        if action:
            sql += " AND ol.action = %s"
            params.append(action)

        cursor.execute(sql, params)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        return rows

    def _create_dataframe(self, rows):
        self.df = pd.DataFrame(rows)
        self.df['created_at'] = pd.to_datetime(self.df['created_at'])

    def _calculate_elapsed_seconds(self):
        self.df['ElapsedSeconds'] = self.df['created_at'].diff().dt.total_seconds().fillna(0).astype(int)

        cols = self.df.columns.tolist()
        if 'created_at' in cols and 'Created_by' in cols:
            created_at_index = cols.index('created_at')
            cols.insert(created_at_index + 1, cols.pop(cols.index('ElapsedSeconds')))
            self.df = self.df[cols]

    def _write_excel(self):
        self.df.to_excel(self.temp_path, index=False)

    def _apply_formatting(self):
        wb = load_workbook(self.temp_path)
        ws = wb.active

        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "ElapsedSeconds":
                self.elapsed_col_idx = col
                break

        if self.elapsed_col_idx:
            self._highlight_cells(ws)
            self._append_summary(ws)

        wb.save(self.temp_path)

    def _highlight_cells(self, ws):
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=self.elapsed_col_idx)
            try:
                value = int(cell.value)
                if value > 60:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = red_fill
                elif value >= 10:
                    cell.fill = red_fill
                elif value >= 5:
                    cell.fill = orange_fill
                else:
                    cell.fill = green_fill
            except:
                continue

    def _append_summary(self, ws):
        created_at_list = self.df['created_at'].tolist()
        elapsed_seconds_list = self.df['ElapsedSeconds'].tolist()

        start_time = created_at_list[0]
        end_time = created_at_list[-1]
        total_seconds = int((end_time - start_time).total_seconds())
        total_h = total_seconds // 3600
        total_m = (total_seconds % 3600) // 60
        total_s = total_seconds % 60
        human_readable = f"{total_h} saat {total_m} dakika {total_s} saniye"

        avg_filtered = [s for s in elapsed_seconds_list if s < 60]
        #avg_seconds = sum(elapsed_seconds_list) / len(elapsed_seconds_list)
        avg_seconds = sum(avg_filtered) / len(avg_filtered) if avg_filtered else 0

        filtered = [s for s in elapsed_seconds_list if s < 10]
        filtered_avg = sum(filtered) / len(filtered) if filtered else 0
        duration = filtered_avg * len(self.df)
        ft_h = int(duration // 3600)
        ft_m = int((duration % 3600) // 60)
        ft_s = int(duration % 60)

        footer_row = ws.max_row + 2
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        ws[f"A{footer_row}"] = "Gercek"
        ws[f"B{footer_row}"] = human_readable
        ws[f"C{footer_row}"] = round(avg_seconds, 6)

        ws[f"A{footer_row + 1}"] = "Filtrelenmiş Ortalama Süre (Kırmızı değerler hariç)"
        ws[f"B{footer_row + 1}"] = f"{ft_h} saat {ft_m} dakika {ft_s} saniye"
        ws[f"C{footer_row + 1}"] = round(filtered_avg, 6)

        for row in range(footer_row, footer_row + 2):
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = thin

    def _save_and_finalize(self):
        os.rename(self.temp_path, self.final_path)

    def _mark_done(self):
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE export_jobs 
            SET status=%s, percent=%s, file_name=%s, file_path=%s 
            WHERE id = %s
        """, ('done', 100, self.job['file_name'], self.final_path, self.job_id))
        conn.commit()
        cursor.close()
        conn.close()
